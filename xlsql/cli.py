"""xlsql converts an Excel .xlsx spreadsheet into a sqlite3 database.

xlsql converts an Excel .xlsx spreadsheet into a sqlite3 database.

Usage: xlsql [OPTIONS] SPREADSHEET

Convert an Excel spreadsheet into a SQLite database.

Args:
    spreadsheet (str): The path to the Excel spreadsheet.

Options:
    --column, -c: A column (or columns) to extract. Can be specified multiple times.
    --database: The name of the database to create. (default: database.db)
    --force: Overwrite an existing database.
    --sheet, -s: A sheet (or sheets) to extract. Can be specified multiple times.
    --verbose, -v: Show verbose output.

Examples:
    xlsql ~/Documents/Example.xlsx
    # Creates: ~/Documents/example.db with all data included in the database.

    xlsql ~/Documents/Example.xlsx --verbose
    # Creates: ~/Documents/example.db, displaying verbose output while running.

    xlsql ~/Documents/Example.xlsx --database /tmp/example.db
    # Creates /tmp/example.db with all data included from the Excel sheet.

    xlsql ~/Documents/Example.xlsx --database /tmp/example.db --force
    # Overwrites the existing db with fresh content from the sheet!

    xlsql example.xlsx -c name -c id -c address -s people
    # Only select the name, id, and address columns from the people sheet.
"""

import sqlite3
from pathlib import Path

import click
import openpyxl

from xlsql.version import VERSION


def normalize(name: str) -> str:
    """
    Normalize a given name by converting it to lowercase, removing
    non-printable characters, replacing hyphens and spaces with underscores.

    Args:
        name (str): The name to be normalized.

    Returns:
        str: The normalized name.
    """
    if name is None:
        return "EMPTY"
    ignored = set("{<([`~!?@#$%^&*,.=:;|])>}")
    replace = set("+- /\\")
    letters = []
    for c in name.lower():
        # Sanitize characters for use in SQL.
        if c.isprintable() and c not in ignored:
            letters.append(c if c not in replace else "_")
        # Remove double underscores.
        if len(letters) > 1:
            if letters[-1] == "_" and letters[-2] == "_":
                letters.pop()

    # Strip leading and trailing underscores.
    while letters and letters[0] == "_":
        letters.pop(0)
    while letters and letters[-1] == "_":
        letters.pop()

    if not letters:
        return "EMPTY"

    # Join the letters and return the normalized name.
    return "".join(letters)


def get_column_names(sheet_name: str, headings: list[str], log: any) -> list[str]:
    """
    Get unique column names for a given sheet.

    Args:
        sheet_name (str): The name of the sheet.
        headings (list[str]): The list of headings.
        log (any): The logging function.

    Returns:
        list[str]: The list of distinct, normalized column names.
    """
    seen = {}
    column_names = []
    for heading in headings:
        normalized = normalize(heading)
        while normalized in seen:
            suffix = seen[normalized]
            seen[normalized] += 1
            normalized = f"{normalized}_{suffix}"
            log(
                f"WARN: duplicate heading in {sheet_name}[{heading}]: renaming to: {normalized}"
            )
        else:
            seen[normalized] = 2
        column_names.append(normalized)
    assert len(column_names) == len(set(column_names))
    return column_names


@click.command()
@click.argument(
    "spreadsheet",
    type=click.Path(exists=True, readable=True),
    default=".",
)
@click.option(
    "--column",
    "-c",
    multiple=True,
    help="A column (or columns) to extract. Can be specified multiple times.",
)
@click.option(
    "--database",
    type=click.Path(exists=False),
    default="database.db",
    help="The name of the database to create.",
)
@click.option(
    "--force",
    is_flag=True,
    type=bool,
    default=False,
    help="Overwrite an existing database.",
)
@click.option(
    "--sheet",
    "-s",
    multiple=True,
    help="A sheet (or sheets) to extract. Can be specified multiple times.",
)
@click.option(
    "--verbose",
    "-v",
    is_flag=True,
    type=bool,
    default=False,
    help="Show verbose output.",
)
@click.option(
    "--version",
    "-V",
    is_flag=True,
    type=bool,
    default=False,
    help="Display the version number.",
)
@click.pass_context
def main(ctx, spreadsheet, column, database, force, sheet, verbose, version) -> None:
    """
    Convert an Excel spreadsheet into a SQLite database.

    SPREADSHEET: A spreadsheet to extract data from.
    """

    # Do nothing if only the version info was requested.
    if version:
        print(VERSION)
        return

    # Display --help output if xlsql was invoked with no arguments, or if a
    # directory was provided instead of a sheet.
    if spreadsheet == "." or Path(spreadsheet).is_dir():
        click.echo(ctx.get_help())
        return

    def log(message: str) -> None:
        if verbose:
            print(message)

    # Ensure that the target database won't be overwritten, or that it's OK to
    # overwrite it.
    existing = Path(database)
    if database and existing.exists() and existing.stat().st_size:
        log(f"Destination database already exists: {database}")
        if force:
            log("Overwriting due to --force flag.")
            existing.unlink()
        else:
            raise click.ClickException(
                f"Cowardly refusing to overwrite existing db: {database} without --force flag"
            )

    try:
        # Load the spreadsheet.
        log(f"Reading contents of speadsheet file: {spreadsheet}")
        workbook = openpyxl.load_workbook(spreadsheet)

        # Create a new SQLite database and connect to it.
        with sqlite3.connect(database) as db:
            log(
                f"Populating {database} using the contents of {len(workbook.sheetnames)} sheets found in {spreadsheet}."
            )

            # Iterate over the sheets in the workbook.
            for sheet_name in workbook.sheetnames:
                # Skip any sheets that were not explicitly requested.
                if (
                    sheet
                    and sheet_name not in sheet
                    and normalize(sheet_name) not in sheet
                ):
                    log(f"Skipping sheet named '{sheet_name}'.")
                    continue

                # Reference the data in the rows of the current sheet.
                rows = workbook[sheet_name].iter_rows(values_only=True)

                # Create a table for each sheet.
                headings = list(next(rows))
                columns = get_column_names(sheet_name, headings, log)
                table_name = normalize(sheet_name)
                log(
                    f"Mapping contents of sheet '{sheet_name}' to table '{table_name}':"
                )

                # Determine whether any columns in this sheet were selected.
                selected = []
                index = 0
                for heading, column_name in zip(headings, columns):
                    if not column or heading in column or column_name in column:
                        selected.append(index)
                        log(f"  {heading} -> {column_name}")
                    index += 1

                # Only create the table if columns were selected.
                if selected:
                    columns = [columns[i] for i in selected]
                    create_table_sql = (
                        f"CREATE TABLE {table_name} ({', '.join(columns)})"
                    )
                    log(f"DB executing SQL: '{create_table_sql};'")
                    db.execute(create_table_sql)
                else:
                    log(
                        f"Skipping table {table_name} because no columns were selected."
                    )
                    continue

                # Insert the rows in batches.
                insert_rows = f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({', '.join(['?'] * len(selected))})"
                cursor = db.cursor()
                batch: list[tuple[str]] = []
                total = [0]

                def insert(row: tuple[str], batch_size: int = 1000) -> None:
                    """
                    Insert a row into the database.

                    Args:
                        row (tuple[str]): The row to be inserted.
                        batch_size (int, optional): The batch size for executing multiple rows at once. Defaults to 1000.

                    Returns:
                        None
                    """
                    flush = False if row else True

                    if row:
                        batch.append(row)

                    if flush or batch and len(batch) >= batch_size:
                        log(f"  ... inserting {len(batch)} rows")
                        cursor.executemany(insert_rows, batch)
                        total[0] += len(batch)
                        batch.clear()
                        if flush:
                            log(f"Writing {total[0]} rows...")
                            db.commit()

                # Insert rows in batches, flushing the final rows.
                log(f"DB executing SQL: '{insert_rows};'")
                for row in rows:
                    if row:
                        if column and selected:
                            insert([row[i] for i in selected])
                        else:
                            insert(row)
                else:
                    insert(None)  # Flush a partial batch.
                    log("DONE!\n")

    finally:
        # Clean up.
        workbook.close()
